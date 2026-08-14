import openpyxl

# Each of the 4 tabs represents ONE statement (not a list of transactions like
# BRCWGS/UCA), so this parser returns a dict-of-dicts, not a row-per-record
# DataFrame: { "COMBINED": {...}, "GEN FUND": {...}, "SEF": {...}, "TRUST FUND": {...} }

TABS = ["COMBINED", "GEN FUND", "SEF", "TRUST FUND"]

LABELS = {
    "total_cash_inflow": "Total Cash Inflow",
    "total_cash_outflow": "Total Cash Outflow",
    "net_cash_operating": "Net Cash from Operating Activities",
    "net_cash_investing": "Net Cash from Investing Activities",
    "net_increase_cash": "Net Increase in Cash",
}


def _find_value(ws, label_text, startswith=False):
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            text = cell.value.strip()
            match = text.startswith(label_text) if startswith else text == label_text
            if not match:
                continue
            for other in row[cell.column - 1:]:
                if isinstance(other.value, (int, float)):
                    return other.value
    return None


def parse_qscf(filepath) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}

    for tab in TABS:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]

        quarter = _find_value(ws, "QUARTER:")
        calendar_year = _find_value(ws, "CALENDAR YEAR:")

        stmt = {key: _find_value(ws, label) for key, label in LABELS.items()}
        stmt["beginning_balance"] = _find_value(ws, "Cash Balance,", startswith=True)

        ending = None
        seen_first = False
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().startswith("Cash Balance,"):
                    if not seen_first:
                        seen_first = True
                        continue
                    for other in row[cell.column - 1:]:
                        if isinstance(other.value, (int, float)):
                            ending = other.value
                            break
                    break
            if ending is not None:
                break
        stmt["ending_balance"] = ending

        stmt["fund_type"] = tab
        stmt["quarter"] = quarter
        stmt["calendar_year"] = calendar_year
        stmt["lgu"] = "City of Cebu"

        result[tab] = stmt

    return result


if __name__ == "__main__":
    from config import QSCF_FILE
    import json
    print(json.dumps(parse_qscf(QSCF_FILE), indent=2, default=str))